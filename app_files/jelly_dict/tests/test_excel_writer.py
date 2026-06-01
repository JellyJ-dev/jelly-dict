from __future__ import annotations

import pickle
import stat
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.core.errors import StorageError
from app.core.models import (
    Example,
    MeaningGroup,
    Sense,
    SubSense,
    VocabularyEntry,
    build_meanings_summary,
    collect_examples_flat,
)
from app.storage.excel_writer import (
    SHEET_NAME,
    append_entry,
    backup_workbook,
    delete_entries_with_backup,
    ensure_workbook,
    find_existing,
    list_entries,
    replace_entry,
    save_with_resolver,
    update_or_append,
)
from app.storage.settings_store import EXCEL_COLUMN_KEYS_DEFAULT


def _entry_apple() -> VocabularyEntry:
    entry = VocabularyEntry(
        language="en",
        word="apple",
        reading="/ˈæp.əl/",
        part_of_speech=["noun"],
        meaning_groups=[
            MeaningGroup(
                pos="noun",
                senses=[
                    Sense(
                        number=1,
                        gloss="a round fruit",
                        sub_senses=[
                            SubSense(
                                gloss="the fruit",
                                examples=[
                                    Example(
                                        source_text="I ate an apple.",
                                        source_text_plain="I ate an apple.",
                                        translation_ko="나는 사과를 먹었다.",
                                    )
                                ],
                            )
                        ],
                    )
                ],
            )
        ],
        memo="first",
        source_url="https://en.dict.naver.com/#/entry/enko/x",
    )
    entry.examples_flat = collect_examples_flat(entry)
    entry.meanings_summary = build_meanings_summary(entry)
    return entry


def test_ensure_workbook_creates_with_header(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    ensure_workbook(path, EXCEL_COLUMN_KEYS_DEFAULT)
    assert path.exists()
    wb = load_workbook(path)
    assert SHEET_NAME in wb.sheetnames
    ws = wb[SHEET_NAME]
    assert ws.cell(row=1, column=1).value == "Language"


def test_append_entry_writes_row(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "en"
    assert ws.cell(row=2, column=2).value == "apple"


def test_find_existing_normalizes_key(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    found = find_existing(path, "en", "apple")
    assert found is not None
    assert found.word == "apple"


def test_update_or_append_replaces_row(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    updated = _entry_apple()
    updated.memo = "second"
    update_or_append(path, updated, EXCEL_COLUMN_KEYS_DEFAULT)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    assert ws.max_row == 2  # not appended
    memo_col = EXCEL_COLUMN_KEYS_DEFAULT.index("memo") + 1
    assert ws.cell(row=2, column=memo_col).value == "second"


def test_list_entries_returns_rows(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)

    from app.storage.excel_writer import list_entries

    entries = list_entries(path)
    assert len(entries) == 1
    assert entries[0].word == "apple"


def test_delete_entries_removes_matching_rows(tmp_path: Path):
    from app.storage.excel_writer import delete_entries, list_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    append_entry(
        path,
        VocabularyEntry(language="en", word="banana"),
        EXCEL_COLUMN_KEYS_DEFAULT,
    )

    removed = delete_entries(path, "en", {"apple"})
    assert removed == 1
    remaining = list_entries(path)
    assert [e.word for e in remaining] == ["banana"]


def test_delete_entries_with_backup_preserves_original_workbook(tmp_path: Path):
    from app.storage.excel_writer import list_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    append_entry(path, VocabularyEntry(language="en", word="banana"), EXCEL_COLUMN_KEYS_DEFAULT)

    outcome = delete_entries_with_backup(path, "en", {"apple"})

    assert outcome.removed == 1
    assert outcome.backup_path is not None
    assert outcome.backup_path.exists()
    assert outcome.backup_path.parent == tmp_path / "Jelly Dict Backups"
    assert [e.word for e in list_entries(path)] == ["banana"]
    assert [e.word for e in list_entries(outcome.backup_path)] == ["apple", "banana"]


def test_delete_entries_with_backup_skips_backup_when_no_rows_match(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)

    outcome = delete_entries_with_backup(path, "en", {"banana"})

    assert outcome.removed == 0
    assert outcome.backup_path is None
    assert not (tmp_path / "Jelly Dict Backups").exists()


def test_backup_workbook_uses_user_visible_backup_folder(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)

    backup = backup_workbook(path, "Manual Snapshot")

    assert backup.exists()
    assert backup.parent.name == "Jelly Dict Backups"
    assert ".manual-snapshot." in backup.name


def test_save_with_resolver_backup_on_overwrite_preserves_previous_row(tmp_path: Path):
    from app.storage.excel_writer import list_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    updated = _entry_apple()
    updated.memo = "second"

    outcome = save_with_resolver(
        path,
        updated,
        EXCEL_COLUMN_KEYS_DEFAULT,
        lambda existing, candidate: ("overwrite", candidate),
        backup_on_overwrite=True,
    )

    action, written = outcome
    assert action == "overwrite"
    assert written.memo == "second"
    assert outcome.backup_path is not None
    assert [entry.memo for entry in list_entries(outcome.backup_path)] == ["first"]
    assert [entry.memo for entry in list_entries(path)] == ["second"]


def test_save_with_resolver_two_value_unpacking_remains_supported(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"

    outcome = save_with_resolver(
        path,
        VocabularyEntry(language="en", word="new"),
        EXCEL_COLUMN_KEYS_DEFAULT,
        lambda existing, candidate: ("create", candidate),
    )

    action, written = outcome
    assert isinstance(outcome, tuple)
    assert len(outcome) == 2
    assert outcome[0] == "create"
    assert outcome[1] == written
    assert outcome == ("create", written)
    assert action == "create"
    assert written.word == "new"


def test_write_outcome_pickle_round_trip_preserves_metadata(tmp_path: Path):
    outcome = save_with_resolver(
        tmp_path / "vocab.xlsx",
        VocabularyEntry(language="en", word="new"),
        EXCEL_COLUMN_KEYS_DEFAULT,
        lambda existing, candidate: ("create", candidate),
    )
    outcome.backup_path = tmp_path / "backup.xlsx"

    restored = pickle.loads(pickle.dumps(outcome))

    assert isinstance(restored, tuple)
    assert restored == outcome
    assert restored.backup_path == tmp_path / "backup.xlsx"


def test_atomic_save_preserves_existing_file_mode(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    path.chmod(0o640)
    updated = _entry_apple()
    updated.memo = "second"

    save_with_resolver(
        path,
        updated,
        EXCEL_COLUMN_KEYS_DEFAULT,
        lambda existing, candidate: ("overwrite", candidate),
        backup_on_overwrite=True,
    )

    assert stat.S_IMODE(path.stat().st_mode) == 0o640


def test_replace_entry_uses_original_key_when_word_changes(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    edited = _entry_apple()
    edited.word = "steer"
    edited.meanings_summary = "조종하다"

    outcome = replace_entry(
        path,
        "en",
        "apple",
        edited,
        EXCEL_COLUMN_KEYS_DEFAULT,
    )

    entries = list_entries(path)
    assert outcome.action == "overwrite"
    assert outcome.backup_path is not None
    assert [entry.word for entry in entries] == ["steer"]
    assert entries[0].meanings_summary == "조종하다"


def test_save_with_resolver_backup_failure_preserves_existing_workbook(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from app.storage import excel_writer
    from app.storage.excel_writer import list_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    updated = _entry_apple()
    updated.memo = "second"

    def fail_copy(src, dst):
        raise OSError("disk full")

    monkeypatch.setattr(excel_writer.shutil, "copy2", fail_copy)

    with pytest.raises(StorageError, match="Excel backup failed"):
        save_with_resolver(
            path,
            updated,
            EXCEL_COLUMN_KEYS_DEFAULT,
            lambda existing, candidate: ("overwrite", candidate),
            backup_on_overwrite=True,
        )

    assert [entry.memo for entry in list_entries(path)] == ["first"]


def test_save_with_resolver_backup_mkdir_failure_is_storage_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from app.storage.excel_writer import list_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    updated = _entry_apple()
    updated.memo = "second"
    original_mkdir = Path.mkdir

    def fail_backup_mkdir(self, *args, **kwargs):
        if self.name == "Jelly Dict Backups":
            raise OSError("mkdir denied")
        return original_mkdir(self, *args, **kwargs)

    monkeypatch.setattr(Path, "mkdir", fail_backup_mkdir)

    with pytest.raises(StorageError, match="Excel backup failed"):
        save_with_resolver(
            path,
            updated,
            EXCEL_COLUMN_KEYS_DEFAULT,
            lambda existing, candidate: ("overwrite", candidate),
            backup_on_overwrite=True,
        )

    assert [entry.memo for entry in list_entries(path)] == ["first"]


def test_save_with_resolver_save_failure_mentions_backup_path(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from app.storage import excel_writer
    from app.storage.excel_writer import list_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    updated = _entry_apple()
    updated.memo = "second"

    def fail_save(wb, target):
        raise StorageError("write failed")

    monkeypatch.setattr(excel_writer, "_save", fail_save)

    with pytest.raises(StorageError) as exc_info:
        save_with_resolver(
            path,
            updated,
            EXCEL_COLUMN_KEYS_DEFAULT,
            lambda existing, candidate: ("overwrite", candidate),
            backup_on_overwrite=True,
        )

    message = str(exc_info.value)
    assert "write failed" in message
    assert "백업 파일:" in message
    assert "Jelly Dict Backups" in message
    assert [entry.memo for entry in list_entries(path)] == ["first"]


def test_save_with_resolver_replace_failure_preserves_existing_workbook(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from app.storage.excel_writer import list_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    updated = _entry_apple()
    updated.memo = "second"

    def fail_replace(self, target):
        raise OSError("replace failed")

    monkeypatch.setattr(Path, "replace", fail_replace)

    with pytest.raises(StorageError) as exc_info:
        save_with_resolver(
            path,
            updated,
            EXCEL_COLUMN_KEYS_DEFAULT,
            lambda existing, candidate: ("overwrite", candidate),
            backup_on_overwrite=True,
        )

    message = str(exc_info.value)
    assert "replace failed" in message
    assert "백업 파일:" in message
    assert [entry.memo for entry in list_entries(path)] == ["first"]


def test_delete_entries_noop_when_keys_missing(tmp_path: Path):
    from app.storage.excel_writer import delete_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)

    assert delete_entries(path, "en", {"banana"}) == 0


def test_update_or_append_appends_when_missing(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    other = VocabularyEntry(language="en", word="banana")
    update_or_append(path, other, EXCEL_COLUMN_KEYS_DEFAULT)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    assert ws.max_row == 3
