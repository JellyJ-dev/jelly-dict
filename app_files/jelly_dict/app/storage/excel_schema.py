from __future__ import annotations

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.core.models import VocabularyEntry, normalize_word_key
from app.storage.excel_serializer import COLUMN_LABELS, COLUMN_WIDTHS, HEADER_FILL, HEADER_FONT


def write_header(ws, columns: list[str]) -> None:
    labels = [COLUMN_LABELS.get(key, key) for key in columns]
    ws.append(labels)
    for idx, key in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(key, 20)
    ws.freeze_panes = "A2"


def style_last_row(ws, columns: list[str]) -> None:
    style_row(ws, ws.max_row, columns)


def style_row(ws, row: int, columns: list[str]) -> None:
    for col_idx, key in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if (
            key == "source_url"
            and isinstance(cell.value, str)
            and cell.value.startswith(("http://", "https://"))
        ):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


def find_row(ws, columns: list[str], entry: VocabularyEntry) -> int | None:
    return find_row_by_key(ws, columns, entry.language, entry.word_key())


def find_row_by_key(
    ws,
    columns: list[str],
    language: str,
    word_key: str,
) -> int | None:
    if "language" not in columns or "word" not in columns:
        return None
    lang_idx = columns.index("language") + 1
    word_idx = columns.index("word") + 1

    for row in range(2, ws.max_row + 1):
        lang_val = ws.cell(row=row, column=lang_idx).value
        word_val = ws.cell(row=row, column=word_idx).value
        if lang_val != language or not isinstance(word_val, str):
            continue
        if normalize_word_key(word_val, language) == word_key:  # type: ignore[arg-type]
            return row
    return None
