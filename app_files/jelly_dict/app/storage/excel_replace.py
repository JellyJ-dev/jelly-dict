from __future__ import annotations

from collections.abc import Callable
from pathlib import Path

from openpyxl.workbook.workbook import Workbook

from app.core.models import VocabularyEntry
from app.storage.excel_outcomes import WriteOutcome
from app.storage.excel_reader import read_header
from app.storage.excel_schema import find_row_by_key, style_last_row, style_row, write_header
from app.storage.excel_serializer import SHEET_NAME, render_cell


def replace_entry(
    path: Path,
    language: str,
    original_word_key: str,
    entry: VocabularyEntry,
    columns: list[str],
    *,
    create_backup: bool,
    append_func: Callable[[Path, VocabularyEntry, list[str]], None],
    backup_func: Callable[[Path, str], Path],
    load_func: Callable[[Path], Workbook],
    save_func: Callable[[Workbook, Path], None],
) -> WriteOutcome:
    if not path.exists():
        append_func(path, entry, columns)
        return WriteOutcome("create", entry)

    wb = load_func(path)
    try:
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            write_header(ws, columns)
        else:
            ws = wb[SHEET_NAME]

        file_columns = read_header(ws) or columns
        if not read_header(ws):
            write_header(ws, file_columns)

        target_row = find_row_by_key(ws, file_columns, language, original_word_key)
        values = [render_cell(entry, key) for key in file_columns]
        backup_path = None
        if target_row is None:
            ws.append(values)
            style_last_row(ws, file_columns)
            action = "create"
        else:
            if create_backup:
                backup_path = backup_func(path, f"edit-{language}")
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=target_row, column=col_idx, value=value)
            style_row(ws, target_row, file_columns)
            action = "overwrite"
        save_func(wb, path)
        return WriteOutcome(action, entry, backup_path)
    finally:
        wb.close()
