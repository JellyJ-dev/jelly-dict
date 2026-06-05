from __future__ import annotations

from collections.abc import Callable
from pathlib import Path

from openpyxl.workbook.workbook import Workbook

from app.core.models import normalize_word_key
from app.storage.excel_backup import backup_workbook
from app.storage.excel_outcomes import DeleteOutcome
from app.storage.excel_reader import read_header
from app.storage.excel_serializer import SHEET_NAME
from app.storage.excel_workbook_io import load_for_write, save_workbook


def delete_entries(path: Path, language: str, word_keys: set[str]) -> int:
    return delete_entries_with_backup(path, language, word_keys, create_backup=False).removed


def delete_entries_with_backup(
    path: Path,
    language: str,
    word_keys: set[str],
    *,
    create_backup: bool = True,
    backup_func: Callable[[Path, str], Path] = backup_workbook,
    load_func: Callable[[Path], Workbook] = load_for_write,
    save_func: Callable[[Workbook, Path], None] = save_workbook,
) -> DeleteOutcome:
    if not path.exists() or not word_keys:
        return DeleteOutcome(0)
    wb = load_func(path)
    try:
        if SHEET_NAME not in wb.sheetnames:
            return DeleteOutcome(0)
        ws = wb[SHEET_NAME]
        columns = read_header(ws)
        if "language" not in columns or "word" not in columns:
            return DeleteOutcome(0)
        lang_idx = columns.index("language") + 1
        word_idx = columns.index("word") + 1

        targets: list[int] = []
        for row in range(2, ws.max_row + 1):
            lang_val = ws.cell(row=row, column=lang_idx).value
            word_val = ws.cell(row=row, column=word_idx).value
            if lang_val != language or not isinstance(word_val, str):
                continue
            key = normalize_word_key(word_val, language)  # type: ignore[arg-type]
            if key in word_keys:
                targets.append(row)
        if not targets:
            return DeleteOutcome(0)
        backup_path = backup_func(path, f"delete-{language}") if create_backup else None
        for row in reversed(targets):
            ws.delete_rows(row, 1)
        save_func(wb, path)
        return DeleteOutcome(len(targets), backup_path)
    finally:
        wb.close()
