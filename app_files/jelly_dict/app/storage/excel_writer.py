"""Write-side helpers for the vocabulary Excel file.

Reader and serializer concerns live in `excel_reader.py` and
`excel_serializer.py` respectively. For backward compatibility this
module re-exports their public symbols so existing callers like
``from app.storage import excel_writer; excel_writer.list_entries(...)``
keep working without import changes.
"""
import shutil
from collections.abc import Callable
from pathlib import Path

from openpyxl import Workbook

from app.core.errors import StorageError
from app.core.models import VocabularyEntry
from app.storage.excel_backup import backup_workbook as _backup_workbook
from app.storage.excel_delete import delete_entries_with_backup as _delete_entries_with_backup
from app.storage.excel_outcomes import DeleteOutcome, WriteOutcome
from app.storage.excel_reader import find_existing, list_entries, read_header
from app.storage.excel_replace import replace_entry as _replace_entry
from app.storage.excel_schema import (
    find_row as _find_row,
    style_last_row as _style_last_row,
    style_row as _style_row,
    write_header as _write_header,
)
from app.storage.excel_serializer import (
    COLUMN_LABELS,
    COLUMN_WIDTHS,
    HEADER_FILL,
    HEADER_FONT,
    SHEET_NAME,
    label_to_key as _label_to_key,
    render_cell as _render_cell,
    row_to_entry as _row_to_entry,
)
from app.storage.excel_workbook_io import (
    load_for_write as _load_for_write,
    raise_with_backup_hint as _raise_with_backup_hint,
    save_workbook,
)

_save = save_workbook
SaveResolver = Callable[[VocabularyEntry | None, VocabularyEntry], tuple[str, VocabularyEntry]]

# Re-exports: keep the existing public surface identical.
__all__ = [
    "SHEET_NAME",
    "COLUMN_LABELS",
    "COLUMN_WIDTHS",
    "HEADER_FILL",
    "HEADER_FONT",
    "DeleteOutcome",
    "WriteOutcome",
    "ensure_workbook",
    "append_entry",
    "update_or_append",
    "backup_workbook",
    "delete_entries",
    "delete_entries_with_backup",
    "replace_entry",
    "save_with_resolver",
    "list_entries",
    "find_existing",
]


def ensure_workbook(path: Path, columns: list[str]) -> None:
    """Create the file with header row if it does not already exist."""
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = SHEET_NAME
        _write_header(ws, columns)
        _save(wb, path)
    finally:
        wb.close()


def append_entry(path: Path, entry: VocabularyEntry, columns: list[str]) -> None:
    """Append a single entry. Creates the file if missing."""
    if not path.exists():
        ensure_workbook(path, columns)

    wb = _load_for_write(path)
    try:
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            _write_header(ws, columns)
        else:
            ws = wb[SHEET_NAME]

        file_columns = read_header(ws) or columns
        if not read_header(ws):
            _write_header(ws, file_columns)

        row_values = [_render_cell(entry, key) for key in file_columns]
        ws.append(row_values)
        _style_last_row(ws, file_columns)
        _save(wb, path)
    finally:
        wb.close()


def update_or_append(path: Path, entry: VocabularyEntry, columns: list[str]) -> None:
    """Replace an existing row with the same (language, word) or append."""
    if not path.exists():
        append_entry(path, entry, columns)
        return

    wb = _load_for_write(path)
    try:
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            _write_header(ws, columns)
        else:
            ws = wb[SHEET_NAME]

        file_columns = read_header(ws) or columns
        target_row = _find_row(ws, file_columns, entry)
        values = [_render_cell(entry, key) for key in file_columns]
        if target_row is None:
            ws.append(values)
            _style_last_row(ws, file_columns)
        else:
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=target_row, column=col_idx, value=value)
            _style_row(ws, target_row, file_columns)
        _save(wb, path)
    finally:
        wb.close()


def backup_workbook(path: Path, reason: str = "manual") -> Path:
    return _backup_workbook(path, reason, copy_func=shutil.copy2)


def delete_entries(path: Path, language: str, word_keys: set[str]) -> int:
    return delete_entries_with_backup(
        path,
        language,
        word_keys,
        create_backup=False,
    ).removed


def delete_entries_with_backup(
    path: Path,
    language: str,
    word_keys: set[str],
    *,
    create_backup: bool = True,
) -> DeleteOutcome:
    return _delete_entries_with_backup(
        path,
        language,
        word_keys,
        create_backup=create_backup,
        backup_func=backup_workbook,
        load_func=_load_for_write,
        save_func=_save,
    )


def replace_entry(
    path: Path,
    language: str,
    original_word_key: str,
    entry: VocabularyEntry,
    columns: list[str],
    *,
    create_backup: bool = True,
) -> WriteOutcome:
    return _replace_entry(
        path,
        language,
        original_word_key,
        entry,
        columns,
        create_backup=create_backup,
        append_func=append_entry,
        backup_func=backup_workbook,
        load_func=_load_for_write,
        save_func=_save,
    )


def save_with_resolver(
    path: Path,
    entry: VocabularyEntry,
    columns: list[str],
    resolver: SaveResolver,
    *,
    backup_on_overwrite: bool = False,
) -> WriteOutcome:
    """Save one entry using an explicit duplicate-resolution action."""
    if not path.exists():
        ensure_workbook(path, columns)

    wb = _load_for_write(path)
    try:
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            _write_header(ws, columns)
        else:
            ws = wb[SHEET_NAME]

        file_columns = read_header(ws) or columns
        if not read_header(ws):
            _write_header(ws, file_columns)

        existing_row = _find_row(ws, file_columns, entry)
        existing_entry: VocabularyEntry | None = None
        if existing_row is not None:
            raw = tuple(
                ws.cell(row=existing_row, column=col_idx).value
                for col_idx in range(1, len(file_columns) + 1)
            )
            existing_entry = _row_to_entry(file_columns, raw)

        action, resolved = resolver(existing_entry, entry)

        if action == "skip":
            # Workbook unchanged — don't bother saving.
            return WriteOutcome(action, existing_entry or entry)

        values = [_render_cell(resolved, key) for key in file_columns]
        backup_path = None
        if action == "overwrite" and existing_row is not None:
            if backup_on_overwrite:
                backup_path = backup_workbook(path, f"overwrite-{entry.language}")
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=existing_row, column=col_idx, value=value)
            _style_row(ws, existing_row, file_columns)
        else:
            # "create" or "append_new" both result in a new row at the end.
            ws.append(values)
            _style_last_row(ws, file_columns)
        try:
            _save(wb, path)
        except StorageError as exc:
            if backup_path is not None:
                _raise_with_backup_hint(exc, backup_path)
            raise
        return WriteOutcome(action, resolved, backup_path)
    finally:
        wb.close()
